import openpyxl
import random

# Create a new Excel workbook
wb = openpyxl.Workbook()

# Select the first sheet
sheet = wb.active

# Define the column headers
headers = ['Roll No', 'Name', 'Math', 'Science', 'English', 'History', 'Geography', 'Total', 'Result']

# Write the headers to the sheet
for i, header in enumerate(headers):
    sheet.cell(row=1, column=i+1).value = header

# Define the student names and roll numbers
names = ['John', 'Jane', 'Bob', 'Alice', 'Mike', 'Emily', 'David', 'Sophia', 'Olivia', 'Ava']
last_names = ['Doe', 'Smith', 'Johnson', 'Brown', 'Davis', 'Chen', 'Lee', 'Patel', 'Martin', 'Kim']

# Generate random marks for each student
for i in range(2, 102):
    roll_no = i - 1
    name = f"{random.choice(names)} {random.choice(last_names)}"
    math_marks = random.randint(25, 95)
    science_marks = random.randint(25, 95)
    english_marks = random.randint(25, 95)
    history_marks = random.randint(25, 95)
    geography_marks = random.randint(25, 95)
    
    # Calculate the total marks
    total_marks = math_marks + science_marks + english_marks + history_marks + geography_marks
    
    # Calculate the result
    if any(marks < 33 for marks in [math_marks, science_marks, english_marks, history_marks, geography_marks]):
        result = "Fail"
    else:
        result = "Pass"
    
    # Write the student data to the sheet
    sheet.cell(row=i, column=1).value = roll_no
    sheet.cell(row=i, column=2).value = name
    sheet.cell(row=i, column=3).value = math_marks
    sheet.cell(row=i, column=4).value = science_marks
    sheet.cell(row=i, column=5).value = english_marks
    sheet.cell(row=i, column=6).value = history_marks
    sheet.cell(row=i, column=7).value = geography_marks
    sheet.cell(row=i, column=8).value = total_marks
    sheet.cell(row=i, column=9).value = result

# Save the workbook
wb.save('New_Student_Marks.xlsx')
print("file created")           